import flickrapi
import pandas as pd
import requests
import os
from io import BytesIO
from PIL import Image
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as xlImage
from datetime import datetime
from dotenv import load_dotenv

# Load the API key from the .env file
load_dotenv()
api_key = os.getenv('API_KEY')
api_secret = os.getenv('API_SECRET')

# Use the API key in your code
flickr = flickrapi.FlickrAPI(api_key, api_secret, format='parsed-json')

# Set username
username = 'ici-et-ailleurs' # 13067705@N08

# Get user's photos
page = 1
per_page = 500
user_photos = []

while True:
    # Get the next page of photos for the user
    photos = flickr.people.getPhotos(user_id=username, page=page, per_page=per_page)

    # Add the photos to the working list
    user_photos += photos['photos']['photo']

    # Check if there are more pages of photos to retrieve
    total_pages = photos['photos']['pages']
    current_page = photos['photos']['page']
    if current_page < total_pages:
        page += 1
    else:
        break
print(f"Retrieved photos from {username}. Total photos: {len(user_photos)}.")

# Create a list of photo IDs from the user's photos
photo_ids = [photo['id'] for photo in user_photos]

# Create an empty pandas DataFrame
df = pd.DataFrame(columns=["Photo ID", "Image", "Creator", "Title", "Date Taken", "Date Uploaded", "License", "Tags"])
print(f"Creating DataFrame to store photo data.")

# Loop over each photo ID
for photo_id in photo_ids:
    # Get photo info
    photo_info = flickr.photos.getInfo(photo_id=photo_id, extras='license')

    # Get thumbnail image URL
    photo_sizes = flickr.photos.getSizes(photo_id=photo_id)
    thumbnail_url = None
    for size in photo_sizes['sizes']['size']:
        if size['label'] == 'Thumbnail':
            thumbnail_url = size['source']
            break
    # Error handling if no image
    if thumbnail_url is None:
        print(f"Error: Thumbnail not found for photo {photo_id}.")

    else:
        # Download and resize the image
        response = requests.get(thumbnail_url)
        img = Image.open(BytesIO(response.content))
        new_size = (75, int(75 * img.size[1] / img.size[0]))
        img = img.resize(new_size)

        # Convert Unix timestamp for upload date to readable format
        date_uploaded = datetime.fromtimestamp(int(photo_info['photo']['dateuploaded']))

        # Dictionary mapping photo license numbers to labels
        license_labels = {
            0: "All Rights Reserved",
            1: "Attribution-NonCommercial-ShareAlike License",
            2: "Attribution-NonCommercial License",
            3: "Attribution-NonCommercial-NoDerivs License",
            4: "Attribution License",
            5: "Attribution-ShareAlike License",
            6: "Attribution-NoDerivs License"
        }

        # Replace license id numbers with the labels
        license = license_labels.get(int(photo_info['photo']['license']), "Unknown")

        # Add row to pandas DataFrame with individual photo data
        row = {
            "Photo ID": photo_info['photo']['id'],
            "Image": thumbnail_url,
            "Creator": photo_info['photo']['owner']['realname'],
            "Title": photo_info['photo']['title']['_content'],
            "Date Taken": photo_info['photo']['dates']['taken'],
            "Date Uploaded": date_uploaded.strftime('%Y-%m-%d %H:%M:%S'),
            "License": license,
            "Tags": "; ".join(tag['_content'] for tag in photo_info['photo']['tags']['tag']),
        }
        df = df._append(row, ignore_index=True)

# Write the DataFrame to an Excel file
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"
print(f"Begin writing DataFrame to Excel file.")

# Write column headers to Excel file
ws.cell(row=1, column=1, value="Photo ID")
ws.cell(row=1, column=2, value="Image")
ws.cell(row=1, column=3, value="Creator")
ws.cell(row=1, column=4, value="Title")
ws.cell(row=1, column=5, value="Date Taken")
ws.cell(row=1, column=6, value="Date Uploaded")
ws.cell(row=1, column=7, value="License")
ws.cell(row=1, column=8, value="Tags")

# Set column width
for col in ws.columns:
    col_letter = col[0].column_letter
    cell_width = 20 # measured in characters
    ws.column_dimensions[col_letter].width = cell_width

for idx, row in df.iterrows():
    # Get URL for each photo
    photo_url = f"https://www.flickr.com/photos/{username}/{row['Photo ID']}"

    row_num = idx + 2
    ws.cell(row=row_num, column=1).value = row['Photo ID']
    ws.cell(row=row_num, column=1).hyperlink = photo_url
    ws.cell(row=row_num, column=1).style = "Hyperlink"

    col_letter = get_column_letter(2)
    if row['Image'] is not None:
        # Add the thumbnail image to the Excel file
        img_data = BytesIO(requests.get(row['Image']).content)
        img = Image.open(img_data)
        img_data.seek(0)
        img_xl = xlImage(img_data)
        img_xl.width = img.size[0]
        img_xl.height = img.size[1]
        ws.column_dimensions[col_letter].width = img_xl.width / 7.5
        ws.row_dimensions[row_num].height = img_xl.height
    ws.add_image(img_xl, f"{col_letter}{row_num}")
    ws.cell(row=row_num, column=3, value=row['Creator'])
    ws.cell(row=row_num, column=4, value=row['Title'])
    ws.cell(row=row_num, column=5, value=row['Date Taken'])
    ws.cell(row=row_num, column=6, value=row['Date Uploaded'])
    ws.cell(row=row_num, column=7, value=row['License'])
    ws.cell(row=row_num, column=8, value=row['Tags'])

    # Print message confirming each photo ID added to file
    # print(f"{row['Photo ID']} written to file.")

# Set save directory path
directory = "data"

# Create /data if it doesn't exist
if not os.path.exists(directory):
    os.makedirs(directory)

# Generate Excel file name
now = datetime.now()
filename = now.strftime("flickrdata_%Y%m%d.xlsx")

# Confirm if file exists in /data
if os.path.isfile(os.path.join(directory, filename)):
    # If file exists add a number to new filename
    i = 1
    while os.path.isfile(os.path.join(directory, f"{filename[:-5]}_{i:03}.xlsx")):
        i += 1
    filename = f"{filename[:-5]}_{i:03}.xlsx"

# Save the file to /data
filepath = os.path.join(directory, filename)
wb.save(filepath)
print(f"Excel file saved as {filename} with {len(df)} rows.")